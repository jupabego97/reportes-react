"""
Rutas de compras v2 — algoritmo con historial completo de 40 meses.
"""
from typing import List, Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.auth.dependencies import get_current_active_user
from app.models.schemas import (
    SugerenciaCompraV2Response,
    ProveedorUrgenciaResponse,
)
from app.services.compras_v2 import ComprasV2Service

router = APIRouter(
    prefix="/api/compras/v2",
    tags=["compras-v2"],
    dependencies=[Depends(get_current_active_user)],
)


@router.get("/sugerencias", response_model=List[SugerenciaCompraV2Response])
async def get_sugerencias_v2(
    proveedor: Optional[str] = Query(None, description="Filtrar por proveedor"),
    db: AsyncSession = Depends(get_db),
):
    """
    Sugerencias de compra calculadas con historial completo (40 meses).
    Incluye velocidad ponderada, ajuste estacional y comparación de precios.
    """
    service = ComprasV2Service(db)
    return await service.get_sugerencias_v2(proveedor=proveedor)


@router.get("/urgencias-proveedor", response_model=List[ProveedorUrgenciaResponse])
async def get_urgencias_proveedor(db: AsyncSession = Depends(get_db)):
    """
    Resumen de urgencias agrupadas por proveedor.
    Usado para las tarjetas de proveedor en la UI de Compras.
    """
    service = ComprasV2Service(db)
    return await service.get_urgencias_por_proveedor()


@router.get("/export-pedido")
async def export_pedido_excel(
    proveedor: str = Query(..., description="Proveedor para el pedido"),
    db: AsyncSession = Depends(get_db),
):
    """
    Exporta el pedido de un proveedor como Excel.
    Solo incluye productos con urgencia urgente/alta/media.
    """
    import io
    from datetime import datetime

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(status_code=500, detail="openpyxl no instalado")

    service = ComprasV2Service(db)
    sugerencias = await service.get_sugerencias_v2(proveedor=proveedor)

    # Filtrar urgencias relevantes
    items = [s for s in sugerencias if s.urgencia in ("urgente", "alta", "media", "baja")]
    if not items:
        items = sugerencias  # si no hay urgentes, incluir todos

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Pedido {proveedor[:20]}"

    # Estilos
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF")
    urgente_fill = PatternFill("solid", fgColor="FFCCCC")
    alta_fill = PatternFill("solid", fgColor="FFE5CC")
    media_fill = PatternFill("solid", fgColor="FFFACC")

    # Título
    ws.merge_cells("A1:I1")
    ws["A1"] = f"Pedido a {proveedor} — {datetime.now().strftime('%d/%m/%Y')}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Encabezados
    headers = [
        "Producto", "Familia", "ABC", "Urgencia",
        "Stock", "Venta/día", "Días stock",
        "Cant. sugerida", "Precio compra", "Costo estimado"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Datos
    total_costo = 0.0
    for row_idx, item in enumerate(items, 3):
        urgencia_label = {
            "urgente": "URGENTE", "alta": "Alta",
            "media": "Media", "baja": "Baja", "ok": "OK"
        }.get(item.urgencia, item.urgencia)

        row_data = [
            item.nombre,
            item.familia or "",
            item.clasificacion_abc,
            urgencia_label,
            item.stock_actual,
            item.velocidad_diaria,
            round(item.dias_stock, 1),
            item.cantidad_sugerida,
            item.precio_ultimo or 0,
            item.costo_estimado,
        ]

        fill = None
        if item.urgencia == "urgente":
            fill = urgente_fill
        elif item.urgencia == "alta":
            fill = alta_fill
        elif item.urgencia == "media":
            fill = media_fill

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            if fill:
                cell.fill = fill
            # Formato moneda para precio y costo
            if col in (9, 10):
                cell.number_format = '"$"#,##0.00'

        total_costo += item.costo_estimado

    # Fila de total
    last_row = len(items) + 3
    ws.cell(row=last_row, column=9, value="TOTAL").font = Font(bold=True)
    total_cell = ws.cell(row=last_row, column=10, value=round(total_costo, 2))
    total_cell.font = Font(bold=True)
    total_cell.number_format = '"$"#,##0.00'

    # Anchos de columna
    col_widths = [40, 20, 6, 10, 8, 10, 10, 14, 14, 15]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    # Devolver como stream
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"pedido_{proveedor.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
